from flask import Flask, render_template, request, send_file, jsonify
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime, timedelta
import os
import glob
from werkzeug.utils import secure_filename

app = Flask(__name__)
# デプロイ環境では一時ファイル用のディレクトリを使用
UPLOAD_FOLDER = os.environ.get('UPLOAD_FOLDER', '.')
app.config['UPLOAD_FOLDER'] = UPLOAD_FOLDER
app.config['MAX_CONTENT_LENGTH'] = 16 * 1024 * 1024  # 16MB max file size
ALLOWED_EXTENSIONS = {'xlsx', 'xls'}

# 法定休憩時間の計算（労働基準法に基づく）
def calculate_required_break_time(work_hours):
    """
    労働時間に応じた法定休憩時間を計算
    - 6時間以下: 休憩不要
    - 6時間超8時間以下: 45分以上
    - 8時間超: 1時間以上
    """
    if work_hours <= 6:
        return 0
    elif work_hours <= 8:
        return 45  # 分
    else:
        return 60  # 分

def parse_time(time_str):
    """時刻文字列をdatetimeオブジェクトに変換"""
    if time_str is None:
        return None
    if isinstance(time_str, datetime):
        return time_str
    try:
        return datetime.strptime(str(time_str), '%Y/%m/%d %H:%M:%S')
    except:
        try:
            return datetime.strptime(str(time_str), '%Y-%m-%d %H:%M:%S')
        except:
            try:
                # 日付なしの時刻形式も試す
                return datetime.strptime(str(time_str), '%H:%M:%S')
            except:
                try:
                    return datetime.strptime(str(time_str), '%H:%M')
                except:
                    return None

def calculate_work_time(start_time, end_time, break_start=None, break_end=None):
    """労働時間を計算（分単位）"""
    if start_time is None or end_time is None:
        return 0
    
    total_minutes = (end_time - start_time).total_seconds() / 60
    
    # 休憩時間を差し引く
    if break_start and break_end:
        break_minutes = (break_end - break_start).total_seconds() / 60
        total_minutes -= break_minutes
    
    return max(0, total_minutes)

def format_time(minutes):
    """分を時間:分の形式に変換"""
    if minutes == 0:
        return "0:00"
    hours = int(minutes // 60)
    mins = int(minutes % 60)
    return f"{hours}:{mins:02d}"

def load_attendance_data(excel_file):
    """エクセルファイルから勤怠データを読み込む"""
    wb = openpyxl.load_workbook(excel_file)
    ws = wb.active
    
    data = []
    headers = []
    
    # ヘッダー行を取得
    for row in ws.iter_rows(min_row=1, max_row=1, values_only=True):
        headers = list(row)
        break
    
    # デバッグ: ヘッダーを確認
    # print("ヘッダー:", headers)
    
    # データ行を取得
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0] is None:  # 空行はスキップ
            continue
        
        row_data = {}
        for i, header in enumerate(headers):
            if header is not None:
                row_data[header] = row[i] if i < len(row) else None
        
        data.append(row_data)
    
    return data, headers

def process_attendance_data(data):
    """勤怠データを処理して個人別・事業別に整理"""
    # 個人別データ
    person_data = {}
    # 事業別データ
    business_data = {}
    
    for record in data:
        # キーを取得
        work_date = record.get('勤務日')
        person_name = record.get('従業員名')
        business_name = record.get('事業所名')
        
        if not work_date or not person_name:
            continue
        
        # 日付を正規化
        if isinstance(work_date, datetime):
            date_key = work_date.date()
        else:
            try:
                date_key = datetime.strptime(str(work_date), '%Y/%m/%d').date()
            except:
                continue
        
        # 時刻をパース
        start_time = parse_time(record.get('出勤時刻'))
        end_time = parse_time(record.get('退勤時刻'))
        break_start = parse_time(record.get('休憩1開始時刻'))
        # 休憩1終了時刻を取得（列名のバリエーションに対応）
        break_end = parse_time(record.get('休憩1終了時刻')) or parse_time(record.get('休憩1復帰時刻'))
        
        # 休憩時間を計算（休憩1終了時刻が休憩復帰時刻の場合）
        break_minutes = 0
        if break_start and break_end:
            break_minutes = (break_end - break_start).total_seconds() / 60
        
        # 労働時間を計算（出勤から退勤までの時間から休憩時間を差し引く）
        # 休憩1終了時刻は休憩復帰時刻として扱う
        work_minutes = calculate_work_time(start_time, end_time, break_start, break_end)
        work_hours = work_minutes / 60
        
        # 法定休憩時間
        required_break = calculate_required_break_time(work_hours)
        
        # 休憩時間が法定内かチェック
        is_break_ok = break_minutes >= required_break if required_break > 0 else True
        
        record_info = {
            'date': date_key,
            'business': business_name,
            'start_time': start_time,
            'end_time': end_time,
            'break_start': break_start,
            'break_end': break_end,
            'work_minutes': work_minutes,
            'work_hours': work_hours,
            'break_minutes': break_minutes,
            'required_break': required_break,
            'is_break_ok': is_break_ok,
            'employee_id': record.get('従業員番号', ''),
        }
        
        # 個人別データに追加
        if person_name not in person_data:
            person_data[person_name] = []
        person_data[person_name].append(record_info)
        
        # 事業別データに追加
        if business_name:
            if business_name not in business_data:
                business_data[business_name] = {}
            if person_name not in business_data[business_name]:
                business_data[business_name][person_name] = []
            business_data[business_name][person_name].append(record_info)
    
    # 日付順にソート
    for person in person_data:
        person_data[person].sort(key=lambda x: x['date'])
    
    for business in business_data:
        for person in business_data[business]:
            business_data[business][person].sort(key=lambda x: x['date'])
    
    return person_data, business_data

def create_output_excel(person_data, business_data, output_file):
    """処理済みデータからエクセルファイルを生成"""
    wb = Workbook()
    
    # デフォルトシートを削除
    if 'Sheet' in wb.sheetnames:
        wb.remove(wb['Sheet'])
    
    # スタイル定義
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    center_align = Alignment(horizontal='center', vertical='center')
    
    # 1. 一覧シートを作成
    create_summary_sheet(wb, person_data, business_data, header_fill, header_font, border, center_align)
    
    # 2. 個人別シートを作成
    for person_name, records in person_data.items():
        create_person_sheet(wb, person_name, records, header_fill, header_font, border, center_align)
    
    # 3. 事業別シートを作成
    for business_name, persons in business_data.items():
        create_business_sheet(wb, business_name, persons, header_fill, header_font, border, center_align)
    
    wb.save(output_file)
    return output_file

def create_summary_sheet(wb, person_data, business_data, header_fill, header_font, border, center_align):
    """一覧シートを作成"""
    ws = wb.create_sheet("一覧", 0)
    
    headers = ['従業員名', '事業所名', '勤務日数', '総労働時間', '平均労働時間', '休憩違反日数']
    ws.append(headers)
    
    # ヘッダー行のスタイル
    for col in range(1, len(headers) + 1):
        cell = ws.cell(1, col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_align
        cell.border = border
    
    # データ行
    row = 2
    for person_name, records in sorted(person_data.items()):
        work_days = len(records)
        total_minutes = sum(r['work_minutes'] for r in records)
        avg_minutes = total_minutes / work_days if work_days > 0 else 0
        violation_days = sum(1 for r in records if not r['is_break_ok'])
        
        # 事業所名を取得（複数事業所の場合は「・」で結合）
        businesses = set()
        for r in records:
            if r['business']:
                businesses.add(r['business'])
        business_name = '・'.join(sorted(businesses)) if businesses else ''
        
        ws.append([
            person_name,
            business_name,
            work_days,
            format_time(total_minutes),
            format_time(avg_minutes),
            violation_days
        ])
        
        # データ行のスタイル
        for col in range(1, len(headers) + 1):
            cell = ws.cell(row, col)
            cell.border = border
            if col >= 3:  # 数値列は中央揃え
                cell.alignment = center_align
        
        row += 1
    
    # 列幅を調整
    ws.column_dimensions['A'].width = 15
    ws.column_dimensions['B'].width = 30  # 事業所名が長くなる可能性があるので広げる
    for col in range(3, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 15

def create_person_sheet(wb, person_name, records, header_fill, header_font, border, center_align):
    """個人別シートを作成"""
    ws = wb.create_sheet(person_name)
    
    headers = ['日付', '出勤時刻', '退勤時刻', '労働時間', '休憩開始', '休憩終了時間', '休憩時間', '法定休憩', '休憩チェック']
    ws.append(headers)
    
    # ヘッダー行のスタイル
    for col in range(1, len(headers) + 1):
        cell = ws.cell(1, col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_align
        cell.border = border
    
    # 休憩違反行用の薄い背景色
    violation_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")  # 薄い黄色
    
    # データ行
    row = 2
    for record in records:
        date_str = record['date'].strftime('%Y/%m/%d')
        start_str = record['start_time'].strftime('%H:%M') if record['start_time'] else ''
        end_str = record['end_time'].strftime('%H:%M') if record['end_time'] else ''
        break_start_str = record['break_start'].strftime('%H:%M') if record['break_start'] else ''
        break_end_str = record['break_end'].strftime('%H:%M') if record['break_end'] else ''
        
        work_time_str = format_time(record['work_minutes'])
        break_time_str = format_time(record['break_minutes'])
        required_break_str = f"{record['required_break']}分" if record['required_break'] > 0 else "不要"
        check_str = "✓" if record['is_break_ok'] else "✗"
        
        ws.append([
            date_str,
            start_str,
            end_str,
            work_time_str,
            break_start_str,
            break_end_str,
            break_time_str,
            required_break_str,
            check_str
        ])
        
        # 休憩違反がある場合は行全体に薄い色をつける
        is_violation = not record['is_break_ok']
        
        # データ行のスタイル
        for col in range(1, len(headers) + 1):
            cell = ws.cell(row, col)
            cell.border = border
            cell.alignment = center_align
            
            # 休憩違反がある場合は行全体に薄い背景色
            if is_violation:
                cell.fill = violation_fill
            
            # 休憩チェック列の色付け（違反の場合は既に背景色がついているので、チェックマークの色は上書きしない）
            if col == 9:  # 休憩チェック列
                if record['is_break_ok']:
                    cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
                else:
                    # 違反の場合は薄い黄色の背景の上に赤い背景を重ねる
                    cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        
        row += 1
    
    # 合計行を追加
    total_minutes = sum(r['work_minutes'] for r in records)
    violation_days = sum(1 for r in records if not r['is_break_ok'])
    
    ws.append(['合計', '', '', format_time(total_minutes), '', '', '', '', f'違反: {violation_days}日'])
    total_row = row
    for col in range(1, len(headers) + 1):
        cell = ws.cell(total_row, col)
        cell.border = border
        cell.font = Font(bold=True)
        if col == 1 or col == 4:
            cell.alignment = center_align
    
    # 列幅を調整
    ws.column_dimensions['A'].width = 12
    for col in range(2, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 12

def create_business_sheet(wb, business_name, persons, header_fill, header_font, border, center_align):
    """事業別シートを作成"""
    ws = wb.create_sheet(f"事業別_{business_name}")
    
    headers = ['従業員名', '勤務日数', '総労働時間', '平均労働時間', '休憩違反日数']
    ws.append(headers)
    
    # ヘッダー行のスタイル
    for col in range(1, len(headers) + 1):
        cell = ws.cell(1, col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_align
        cell.border = border
    
    # データ行
    row = 2
    for person_name, records in sorted(persons.items()):
        work_days = len(records)
        total_minutes = sum(r['work_minutes'] for r in records)
        avg_minutes = total_minutes / work_days if work_days > 0 else 0
        violation_days = sum(1 for r in records if not r['is_break_ok'])
        
        ws.append([
            person_name,
            work_days,
            format_time(total_minutes),
            format_time(avg_minutes),
            violation_days
        ])
        
        # データ行のスタイル
        for col in range(1, len(headers) + 1):
            cell = ws.cell(row, col)
            cell.border = border
            if col >= 2:  # 数値列は中央揃え
                cell.alignment = center_align
        
        row += 1
    
    # 列幅を調整
    ws.column_dimensions['A'].width = 15
    for col in range(2, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 15

def create_business_excel(business_name, persons, output_file):
    """事業所別のエクセルファイルを生成"""
    wb = Workbook()
    
    # デフォルトシートを削除
    if 'Sheet' in wb.sheetnames:
        wb.remove(wb['Sheet'])
    
    # スタイル定義
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    center_align = Alignment(horizontal='center', vertical='center')
    
    # 一覧シートを作成
    ws = wb.create_sheet("一覧", 0)
    headers = ['従業員名', '勤務日数', '総労働時間', '平均労働時間', '休憩違反日数']
    ws.append(headers)
    
    # ヘッダー行のスタイル
    for col in range(1, len(headers) + 1):
        cell = ws.cell(1, col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_align
        cell.border = border
    
    # データ行
    row = 2
    for person_name, records in sorted(persons.items()):
        work_days = len(records)
        total_minutes = sum(r['work_minutes'] for r in records)
        avg_minutes = total_minutes / work_days if work_days > 0 else 0
        violation_days = sum(1 for r in records if not r['is_break_ok'])
        
        ws.append([
            person_name,
            work_days,
            format_time(total_minutes),
            format_time(avg_minutes),
            violation_days
        ])
        
        # データ行のスタイル
        for col in range(1, len(headers) + 1):
            cell = ws.cell(row, col)
            cell.border = border
            if col >= 2:  # 数値列は中央揃え
                cell.alignment = center_align
        
        row += 1
    
    # 列幅を調整
    ws.column_dimensions['A'].width = 15
    for col in range(2, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 15
    
    # 個人別シートを作成
    for person_name, records in sorted(persons.items()):
        create_person_sheet(wb, person_name, records, header_fill, header_font, border, center_align)
    
    wb.save(output_file)
    return output_file

@app.route('/')
def index():
    """メインページ"""
    # フォルダ内のエクセルファイルを取得
    excel_files = glob.glob('*.xlsx') + glob.glob('*.xls')
    excel_files = [f for f in excel_files if not f.startswith('~$')]  # 一時ファイルを除外
    
    return render_template('index.html', excel_files=excel_files)

def allowed_file(filename):
    """ファイル拡張子をチェック"""
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS

@app.route('/process', methods=['POST'])
def process():
    """エクセルファイルを処理"""
    try:
        excel_file = None
        uploaded_file = None
        
        # ファイルアップロードをチェック
        if 'file' in request.files:
            file = request.files['file']
            if file and file.filename and allowed_file(file.filename):
                # アップロードされたファイルを保存
                filename = secure_filename(file.filename)
                timestamp = datetime.now().strftime('%Y%m%d%H%M%S')
                excel_file = f'uploaded_{timestamp}_{filename}'
                file.save(excel_file)
                uploaded_file = excel_file
        
        # フォームからファイル名を取得（既存ファイルの場合）
        if not excel_file:
            excel_file = request.form.get('excel_file')
        
        if not excel_file or not os.path.exists(excel_file):
            return jsonify({'error': 'ファイルが見つかりません'}), 400
        
        # データを読み込む
        data, headers = load_attendance_data(excel_file)
        
        # データを処理
        person_data, business_data = process_attendance_data(data)
        
        # 出力ファイル名を生成
        timestamp = datetime.now().strftime('%Y%m%d%H%M%S')
        output_file = f'勤怠集計_{timestamp}.xlsx'
        
        # エクセルファイルを生成
        create_output_excel(person_data, business_data, output_file)
        
        # 事業所別ファイルを生成
        business_files = {}
        for business_name, persons in business_data.items():
            # 事業所名からファイル名に使えない文字を削除
            safe_business_name = business_name.replace('/', '_').replace('\\', '_').replace(':', '_')
            business_output_file = f'勤怠集計_{safe_business_name}_{timestamp}.xlsx'
            create_business_excel(business_name, persons, business_output_file)
            business_files[business_name] = business_output_file
        
        # アップロードされたファイルを削除（一時ファイル）
        if uploaded_file and os.path.exists(uploaded_file):
            try:
                os.remove(uploaded_file)
            except:
                pass
        
        return jsonify({
            'success': True,
            'output_file': output_file,
            'person_count': len(person_data),
            'business_count': len(business_data),
            'business_files': business_files
        })
    
    except Exception as e:
        # エラー時もアップロードファイルを削除
        if uploaded_file and os.path.exists(uploaded_file):
            try:
                os.remove(uploaded_file)
            except:
                pass
        return jsonify({'error': str(e)}), 500

@app.route('/download/<filename>')
def download(filename):
    """生成されたファイルをダウンロード"""
    return send_file(filename, as_attachment=True)

if __name__ == '__main__':
    port = int(os.environ.get('PORT', 5000))
    app.run(host='0.0.0.0', port=port, debug=False)

